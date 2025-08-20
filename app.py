from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, g, send_file
from flask_bcrypt import Bcrypt
import os
import sys
import sqlite3
from datetime import datetime
import csv
from io import StringIO, BytesIO
try:
    import openpyxl
except Exception:
    openpyxl = None

from ml_utils import churn_summary, ensure_ready

# Resolve paths robustly for frozen executables and Windows data directory
_IS_FROZEN = getattr(sys, 'frozen', False)
_BASE_DIR_EXEC = os.path.dirname(getattr(sys, 'executable', sys.argv[0])) if _IS_FROZEN else os.path.dirname(__file__)
_MEIPASS = getattr(sys, '_MEIPASS', None)

# Base directory for bundled assets (templates, models)
BASE_DIR = _MEIPASS or _BASE_DIR_EXEC

# Database path: use a user-writable directory on Windows
if os.name == 'nt':
    appdata = os.getenv('LOCALAPPDATA') or os.path.expanduser('~')
    data_dir = os.path.join(appdata, 'NIVARA')
    os.makedirs(data_dir, exist_ok=True)
    DB_PATH = os.path.join(data_dir, 'crm.db')
else:
    DB_PATH = os.path.join(BASE_DIR, 'crm.db')

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))
# Use fixed secret key from env in production for stable sessions; fallback to random in dev
app.secret_key = os.getenv('SECRET_KEY') or os.urandom(24)
bcrypt = Bcrypt(app)

# ----------------------
# Database helpers
# ----------------------

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = get_db()

    # Users
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    # Customers
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            mobile TEXT,
            phone TEXT,
            address TEXT,
            city TEXT,
            pincode TEXT,
            state TEXT,
            country TEXT,
            notes TEXT,
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    # Leads
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            mobile TEXT,
            phone TEXT,
            source TEXT,
            status TEXT DEFAULT 'new',
            notes TEXT,
            customer_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(customer_id) REFERENCES customers(id)
        );
        """
    )

    # Opportunities
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS opportunities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            amount REAL,
            stage TEXT DEFAULT 'prospecting',
            status TEXT DEFAULT 'open',
            customer_id INTEGER,
            lead_id INTEGER,
            notes TEXT,
            close_date TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(customer_id) REFERENCES customers(id),
            FOREIGN KEY(lead_id) REFERENCES leads(id)
        );
        """
    )

    # Seed default user if table is empty or the demo email doesn't exist
    existing = db.execute('SELECT id FROM users WHERE email = ?', ('user@example.com',)).fetchone()
    if not existing:
        db.execute(
            'INSERT INTO users (name, email, password_hash, created_at) VALUES (?, ?, ?, ?)',
            (
                'Test User',
                'user@example.com',
                bcrypt.generate_password_hash('password123').decode('utf-8'),
                datetime.utcnow().isoformat(),
            ),
        )

    db.commit()

# ----------------------
# Auth helpers
# ----------------------

def login_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if 'user_email' not in session:
            return redirect(url_for('login'))
        return view_func(*args, **kwargs)
    return wrapped


def get_user_by_email(email: str):
    if not email:
        return None
    db = get_db()
    return db.execute('SELECT * FROM users WHERE email = ?', (email.lower(),)).fetchone()

# ----------------------
# Routes - Auth + Dashboard
# ----------------------

def _days_between(d1_str, d2):
    try:
        if not d1_str:
            return None
        d1 = datetime.fromisoformat(d1_str)
        return (d2 - d1).days
    except Exception:
        return None


def _lead_score(row):
    # Simple heuristic score [0..100]
    score = 0
    # Contact completeness
    if row.get('email'): score += 10
    if row.get('mobile'): score += 8
    if row.get('phone'): score += 5
    # Source
    src = (row.get('source') or '').lower()
    score += {'referral': 15, 'web': 10, 'campaign': 8, 'ad': 5}.get(src, 3 if src else 0)
    # Status
    status = (row.get('status') or '').lower()
    score += {'new': 5, 'contacted': 12, 'qualified': 22, 'lost': -15}.get(status, 0)
    # Recency: newer leads rank higher
    now = datetime.utcnow()
    age_days = _days_between(row.get('created_at'), now)
    if age_days is not None:
        if age_days <= 3:
            score += 12
        elif age_days <= 7:
            score += 8
        elif age_days <= 30:
            score += 3
        else:
            score += 0
    # Cap and floor
    return max(0, min(100, score))


def _opp_win_probability(row):
    # Heuristic win probability [0..1]
    stage = (row.get('stage') or '').lower()
    base = {
        'prospecting': 0.2,
        'qualification': 0.35,
        'proposal': 0.55,
        'negotiation': 0.7,
        'won': 0.98,
        'lost': 0.02,
    }.get(stage, 0.3)
    prob = base
    # Amount scaling: smaller deals slightly easier
    amt = row.get('amount') or 0
    try:
        amt = float(amt or 0)
    except Exception:
        amt = 0.0
    if amt:
        if amt < 1000: prob += 0.05
        elif amt < 5000: prob += 0.03
        elif amt > 20000: prob -= 0.03
    # Close date proximity
    now = datetime.utcnow()
    cd = row.get('close_date')
    try:
        if cd:
            cd_dt = datetime.fromisoformat(cd)
            days_to_close = (cd_dt - now).days
            if days_to_close >= 0 and days_to_close <= 14:
                prob += 0.07
            elif days_to_close < 0 and (row.get('status') or '').lower() == 'open':
                prob -= 0.12
    except Exception:
        pass
    # Links
    if row.get('customer_id'): prob += 0.04
    if row.get('lead_id'): prob += 0.02
    # Status override
    status = (row.get('status') or '').lower()
    if status == 'closed':
        prob = 0.0 if stage == 'lost' else min(1.0, prob)
    # Clamp
    return max(0.0, min(1.0, prob))


def _compute_analytics():
    db = get_db()
    # Counts
    cust_count = db.execute('SELECT COUNT(*) as c FROM customers').fetchone()['c']
    lead_count = db.execute('SELECT COUNT(*) as c FROM leads').fetchone()['c']
    opp_count = db.execute('SELECT COUNT(*) as c FROM opportunities').fetchone()['c']

    # Leads by status
    lead_by_status = db.execute('SELECT status, COUNT(*) as c FROM leads GROUP BY status').fetchall()
    lead_by_status = { (r['status'] or 'unknown'): r['c'] for r in lead_by_status }

    # Opportunities by stage and totals
    opp_by_stage = db.execute('SELECT stage, COUNT(*) as c, SUM(amount) as total FROM opportunities GROUP BY stage').fetchall()
    opp_by_stage = [ {'stage': (r['stage'] or 'unknown'), 'count': r['c'], 'total': float(r['total'] or 0)} for r in opp_by_stage ]
    pipeline_total = sum(o['total'] for o in opp_by_stage)

    # Win/loss stats
    won = db.execute("SELECT COUNT(*) as c, SUM(amount) as total FROM opportunities WHERE stage='won'").fetchone()
    lost = db.execute("SELECT COUNT(*) as c FROM opportunities WHERE stage='lost'").fetchone()
    won_count = won['c'] or 0
    lost_count = lost['c'] or 0
    win_rate = (won_count / (won_count + lost_count)) if (won_count + lost_count) else 0
    won_total = float(won['total'] or 0)

    # Top leads by score
    leads = db.execute('SELECT * FROM leads ORDER BY updated_at DESC LIMIT 200').fetchall()
    scored_leads = []
    for l in leads:
        d = dict(l)
        d['score'] = _lead_score(d)
        scored_leads.append(d)
    top_leads = sorted(scored_leads, key=lambda x: x['score'], reverse=True)[:5]

    # Top opportunities by probability
    opps = db.execute('SELECT * FROM opportunities ORDER BY updated_at DESC LIMIT 200').fetchall()
    scored_opps = []
    for o in opps:
        d = dict(o)
        d['prob'] = _opp_win_probability(d)
        scored_opps.append(d)
    top_opps = sorted(scored_opps, key=lambda x: x['prob'], reverse=True)[:5]

    # Ensure churn model is ready and compute churn KPIs
    try:
        ensure_ready(db)
        churn = churn_summary(db)
    except Exception:
        churn = {
            'avg_risk': 0.0,
            'high_risk_percent': 0.0,
            'expected_churn_amount': 0.0,
            'top_customers': [],
            'total_open_pipeline': 0.0,
        }

    # Suggestions / next-best actions
    suggestions = []
    # Follow up new leads
    for l in scored_leads:
        if (l.get('status') or '').lower() in ('new', 'contacted') and l['score'] >= 20:
            suggestions.append(f"Follow up with lead '{l['name']}' (score {l['score']})")
            if len(suggestions) >= 3: break
    # Opportunities near close date with mid probability
    now = datetime.utcnow()
    for o in scored_opps:
        cd = o.get('close_date')
        try:
            if cd:
                days = (datetime.fromisoformat(cd) - now).days
                if 0 <= days <= 14 and 0.4 <= o['prob'] <= 0.75:
                    suggestions.append(f"Nudge opportunity '{o['title']}' (p(win) {o['prob']:.0%}, closes in {days}d)")
        except Exception:
            pass
        if len(suggestions) >= 6: break

    return {
        'counts': {'customers': cust_count, 'leads': lead_count, 'opportunities': opp_count},
        'lead_by_status': lead_by_status,
        'opp_by_stage': opp_by_stage,
        'pipeline_total': pipeline_total,
        'win_rate': win_rate,
        'won_total': won_total,
        'top_leads': top_leads,
        'top_opps': top_opps,
        'suggestions': suggestions,
        'churn': churn,
    }


@app.route('/')
@login_required
def home():
    user = get_user_by_email(session.get('user_email'))
    name = user['name'] if user else 'User'
    analytics = _compute_analytics()
    return render_template('dashboard.html', name=name, analytics=analytics)

@app.route('/api/analytics')
@login_required
def api_analytics():
    return jsonify(_compute_analytics())

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        password = request.form.get('password') or ''
        user = get_user_by_email(email)
        if user and bcrypt.check_password_hash(user['password_hash'], password):
            session['user_email'] = email
            session['user_id'] = user['id']
            flash('You were successfully logged in')
            return redirect(url_for('home'))
        else:
            error = 'Invalid credentials. Please try again.'
    return render_template('login.html', error=error)

@app.route('/register', methods=['GET', 'POST'])
def register():
    error = None
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        email = (request.form.get('email') or '').strip().lower()
        password = request.form.get('password') or ''
        confirm = request.form.get('confirm') or ''

        if not name or not email or not password:
            error = 'Name, email, and password are required.'
        elif password != confirm:
            error = 'Passwords do not match.'
        else:
            db = get_db()
            exists = db.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()
            if exists:
                error = 'An account with this email already exists.'
            else:
                db.execute(
                    'INSERT INTO users (name, email, password_hash, created_at) VALUES (?, ?, ?, ?)',
                    (name, email, bcrypt.generate_password_hash(password).decode('utf-8'), datetime.utcnow().isoformat()),
                )
                db.commit()
                flash('Account created. You can now log in.')
                return redirect(url_for('login'))

    return render_template('register.html', error=error)

@app.route('/logout')
def logout():
    session.pop('user_email', None)
    session.pop('user_id', None)
    flash('You have been logged out')
    return redirect(url_for('login'))

# ----------------------
# Customers - UI
# ----------------------

@app.route('/customers')
@login_required
def customers_list():
    db = get_db()
    rows = db.execute("SELECT * FROM customers ORDER BY updated_at DESC").fetchall()
    return render_template('customers/list.html', customers=rows)

@app.route('/customers/export.<fmt>')
@login_required
def customers_export(fmt):
    db = get_db()
    rows = db.execute('SELECT * FROM customers ORDER BY id').fetchall()
    headers = ['id','name','email','mobile','phone','address','city','pincode','state','country','notes','status','created_at','updated_at']
    if fmt == 'csv':
        si = StringIO()
        writer = csv.writer(si)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([r[h] for h in headers])
        si.seek(0)
        bio = BytesIO(si.getvalue().encode('utf-8'))
        bio.seek(0)
        return send_file(bio, mimetype='text/csv', as_attachment=True, download_name='customers.csv')
    elif fmt in ('xlsx','xls'):
        if not openpyxl:
            return 'openpyxl not installed', 500
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Customers'
        ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='customers.xlsx')
    else:
        return 'Unsupported format', 400

@app.route('/customers/sample.csv')
@login_required
def customers_sample_csv():
    headers = ['name','email','mobile','phone','address','city','pincode','state','country','notes','status']
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(headers)
    writer.writerow(['Acme Corp','info@acme.com','9876543210','','123 Lane','Mumbai','400001','MH','India','Important client','active'])
    si.seek(0)
    bio = BytesIO(si.getvalue().encode('utf-8'))
    bio.seek(0)
    return send_file(bio, mimetype='text/csv', as_attachment=True, download_name='customers_sample.csv')

@app.route('/customers/import', methods=['POST'])
@login_required
def customers_import():
    file = request.files.get('file')
    if not file:
        flash('No file uploaded')
        return redirect(url_for('customers_list'))
    filename = (file.filename or '').lower()
    count = 0
    db = get_db()
    now = datetime.utcnow().isoformat()

    def upsert(row):
        nonlocal count
        name = (row.get('name') or '').strip()
        if not name:
            return
        db.execute('''INSERT INTO customers (name, email, mobile, phone, address, city, pincode, state, country, notes, status, created_at, updated_at)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                   (name, row.get('email'), row.get('mobile'), row.get('phone'), row.get('address'), row.get('city'), row.get('pincode'), row.get('state'), row.get('country'), row.get('notes'), (row.get('status') or 'active'), now, now))
        count += 1

    if filename.endswith('.csv'):
        data = file.read().decode('utf-8')
        reader = csv.DictReader(StringIO(data))
        for row in reader:
            upsert(row)
    elif filename.endswith(('.xlsx', '.xls')):
        if not openpyxl:
            flash('openpyxl not installed on server')
            return redirect(url_for('customers_list'))
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            upsert({headers[i]: row[i] for i in range(len(headers))})
    else:
        flash('Unsupported file format. Use CSV or XLSX.')
        return redirect(url_for('customers_list'))

    db.commit()
    flash(f'Imported {count} customers')
    return redirect(url_for('customers_list'))

@app.route('/customers/new', methods=['GET', 'POST'])
@login_required
def customers_new():
    if request.method == 'POST':
        data = {
            'name': request.form.get('name', '').strip(),
            'email': request.form.get('email'),
            'mobile': request.form.get('mobile'),
            'phone': request.form.get('phone'),
            'address': request.form.get('address'),
            'city': request.form.get('city'),
            'pincode': request.form.get('pincode'),
            'state': request.form.get('state'),
            'country': request.form.get('country'),
            'notes': request.form.get('notes'),
            'status': request.form.get('status', 'active')
        }
        if not data['name']:
            flash('Name is required')
            return redirect(url_for('customers_new'))
        db = get_db()
        db.execute(
            '''INSERT INTO customers (name, email, mobile, phone, address, city, pincode, state, country, notes, status, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (data['name'], data['email'], data['mobile'], data['phone'], data['address'], data['city'], data['pincode'], data['state'], data['country'], data['notes'], data['status'], datetime.utcnow().isoformat(), datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Customer created')
        return redirect(url_for('customers_list'))
    return render_template('customers/form.html', customer=None)

@app.route('/customers/<int:customer_id>/edit', methods=['GET', 'POST'])
@login_required
def customers_edit(customer_id):
    db = get_db()
    row = db.execute('SELECT * FROM customers WHERE id = ?', (customer_id,)).fetchone()
    if not row:
        flash('Customer not found')
        return redirect(url_for('customers_list'))
    if request.method == 'POST':
        data = {
            'name': request.form.get('name', '').strip(),
            'email': request.form.get('email'),
            'mobile': request.form.get('mobile'),
            'phone': request.form.get('phone'),
            'address': request.form.get('address'),
            'city': request.form.get('city'),
            'pincode': request.form.get('pincode'),
            'state': request.form.get('state'),
            'country': request.form.get('country'),
            'notes': request.form.get('notes'),
            'status': request.form.get('status', row['status'])
        }
        if not data['name']:
            flash('Name is required')
            return redirect(url_for('customers_edit', customer_id=customer_id))
        db.execute(
            '''UPDATE customers SET name=?, email=?, mobile=?, phone=?, address=?, city=?, pincode=?, state=?, country=?, notes=?, status=?, updated_at=? WHERE id=?''',
            (data['name'], data['email'], data['mobile'], data['phone'], data['address'], data['city'], data['pincode'], data['state'], data['country'], data['notes'], data['status'], datetime.utcnow().isoformat(), customer_id)
        )
        db.commit()
        flash('Customer updated')
        return redirect(url_for('customers_list'))
    return render_template('customers/form.html', customer=row)

@app.route('/customers/<int:customer_id>/delete', methods=['POST'])
@login_required
def customers_delete(customer_id):
    db = get_db()
    db.execute('DELETE FROM customers WHERE id = ?', (customer_id,))
    db.commit()
    flash('Customer deleted')
    return redirect(url_for('customers_list'))

# ----------------------
# Leads - UI
# ----------------------

@app.route('/leads')
@login_required
def leads_list():
    db = get_db()
    rows = db.execute('SELECT * FROM leads ORDER BY updated_at DESC').fetchall()
    return render_template('leads/list.html', leads=rows)

@app.route('/leads/export.<fmt>')
@login_required
def leads_export(fmt):
    db = get_db()
    rows = db.execute('SELECT * FROM leads ORDER BY id').fetchall()
    headers = ['id','name','email','mobile','phone','source','status','notes','customer_id','created_at','updated_at']
    if fmt == 'csv':
        si = StringIO()
        writer = csv.writer(si)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([r[h] for h in headers])
        si.seek(0)
        bio = BytesIO(si.getvalue().encode('utf-8'))
        bio.seek(0)
        return send_file(bio, mimetype='text/csv', as_attachment=True, download_name='leads.csv')
    elif fmt in ('xlsx','xls'):
        if not openpyxl:
            return 'openpyxl not installed', 500
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leads'
        ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='leads.xlsx')
    else:
        return 'Unsupported format', 400

@app.route('/leads/import', methods=['POST'])
@login_required
def leads_import():
    file = request.files.get('file')
    if not file:
        flash('No file uploaded')
        return redirect(url_for('leads_list'))
    filename = (file.filename or '').lower()
    count = 0
    db = get_db()
    now = datetime.utcnow().isoformat()

    def insert_from(row):
        nonlocal count
        name = (row.get('name') or '').strip()
        if not name:
            return
        db.execute('''INSERT INTO leads (name, email, mobile, phone, source, status, notes, customer_id, created_at, updated_at)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                   (name, row.get('email'), row.get('mobile'), row.get('phone'), row.get('source'), (row.get('status') or 'new'), row.get('notes'), row.get('customer_id') or None, now, now))
        count += 1

    if filename.endswith('.csv'):
        data = file.read().decode('utf-8')
        reader = csv.DictReader(StringIO(data))
        for row in reader:
            insert_from(row)
    elif filename.endswith(('.xlsx', '.xls')):
        if not openpyxl:
            flash('openpyxl not installed on server')
            return redirect(url_for('leads_list'))
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            insert_from({headers[i]: row[i] for i in range(len(headers))})
    else:
        flash('Unsupported file format. Use CSV or XLSX.')
        return redirect(url_for('leads_list'))

    db.commit()
    flash(f'Imported {count} leads')
    return redirect(url_for('leads_list'))

@app.route('/leads/new', methods=['GET', 'POST'])
@login_required
def leads_new():
    db = get_db()
    customers = db.execute('SELECT id, name FROM customers ORDER BY name').fetchall()
    if request.method == 'POST':
        data = {
            'name': request.form.get('name', '').strip(),
            'email': request.form.get('email'),
            'mobile': request.form.get('mobile'),
            'phone': request.form.get('phone'),
            'source': request.form.get('source'),
            'status': request.form.get('status', 'new'),
            'notes': request.form.get('notes'),
            'customer_id': request.form.get('customer_id') or None
        }
        if not data['name']:
            flash('Name is required')
            return redirect(url_for('leads_new'))
        db.execute(
            '''INSERT INTO leads (name, email, mobile, phone, source, status, notes, customer_id, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (data['name'], data['email'], data['mobile'], data['phone'], data['source'], data['status'], data['notes'], data['customer_id'], datetime.utcnow().isoformat(), datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Lead created')
        return redirect(url_for('leads_list'))
    return render_template('leads/form.html', lead=None, customers=customers)

@app.route('/leads/<int:lead_id>/edit', methods=['GET', 'POST'])
@login_required
def leads_edit(lead_id):
    db = get_db()
    lead = db.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
    customers = db.execute('SELECT id, name FROM customers ORDER BY name').fetchall()
    if not lead:
        flash('Lead not found')
        return redirect(url_for('leads_list'))
    if request.method == 'POST':
        data = {
            'name': request.form.get('name', '').strip(),
            'email': request.form.get('email'),
            'mobile': request.form.get('mobile'),
            'phone': request.form.get('phone'),
            'source': request.form.get('source'),
            'status': request.form.get('status', lead['status']),
            'notes': request.form.get('notes'),
            'customer_id': request.form.get('customer_id') or None
        }
        if not data['name']:
            flash('Name is required')
            return redirect(url_for('leads_edit', lead_id=lead_id))
        db.execute(
            '''UPDATE leads SET name=?, email=?, mobile=?, phone=?, source=?, status=?, notes=?, customer_id=?, updated_at=? WHERE id=?''',
            (data['name'], data['email'], data['mobile'], data['phone'], data['source'], data['status'], data['notes'], data['customer_id'], datetime.utcnow().isoformat(), lead_id)
        )
        db.commit()
        flash('Lead updated')
        return redirect(url_for('leads_list'))
    return render_template('leads/form.html', lead=lead, customers=customers)

@app.route('/leads/<int:lead_id>/delete', methods=['POST'])
@login_required
def leads_delete(lead_id):
    db = get_db()
    db.execute('DELETE FROM leads WHERE id = ?', (lead_id,))
    db.commit()
    flash('Lead deleted')
    return redirect(url_for('leads_list'))

# ----------------------
# Opportunities - UI
# ----------------------

@app.route('/opportunities')
@login_required
def opportunities_list():
    db = get_db()
    rows = db.execute('SELECT * FROM opportunities ORDER BY updated_at DESC').fetchall()
    return render_template('opportunities/list.html', opportunities=rows)

@app.route('/opportunities/export.<fmt>')
@login_required
def opportunities_export(fmt):
    db = get_db()
    rows = db.execute('SELECT * FROM opportunities ORDER BY id').fetchall()
    headers = ['id','title','amount','stage','status','customer_id','lead_id','notes','close_date','created_at','updated_at']
    if fmt == 'csv':
        si = StringIO()
        writer = csv.writer(si)
        writer.writerow(headers)
        for r in rows:
            # amount is in INR; keep numeric raw
            writer.writerow([r[h] for h in headers])
        si.seek(0)
        bio = BytesIO(si.getvalue().encode('utf-8'))
        bio.seek(0)
        return send_file(bio, mimetype='text/csv', as_attachment=True, download_name='opportunities.csv')
    elif fmt in ('xlsx','xls'):
        if not openpyxl:
            return 'openpyxl not installed', 500
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Opportunities'
        ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='opportunities.xlsx')
    else:
        return 'Unsupported format', 400

@app.route('/opportunities/sample.csv')
@login_required
def opportunities_sample_csv():
    headers = ['title','amount','stage','status','customer_id','lead_id','notes','close_date']
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(headers)
    writer.writerow(['Website redesign', '15000', 'proposal', 'open', '', '', 'Priority client', '2025-09-30'])
    si.seek(0)
    bio = BytesIO(si.getvalue().encode('utf-8'))
    bio.seek(0)
    return send_file(bio, mimetype='text/csv', as_attachment=True, download_name='opportunities_sample.csv')

@app.route('/opportunities/import', methods=['POST'])
@login_required
def opportunities_import():
    file = request.files.get('file')
    if not file:
        flash('No file uploaded')
        return redirect(url_for('opportunities_list'))
    filename = (file.filename or '').lower()
    count = 0
    db = get_db()
    now = datetime.utcnow().isoformat()

    def to_float(val):
        # Parse INR values like "₹ 10,000.50" or plain numbers to float
        if val is None:
            return None
        s = str(val)
        s = s.replace('₹', '').replace(',', '').strip()
        try:
            return float(s) if s else None
        except Exception:
            return None

    def insert_from(row):
        nonlocal count
        title = (row.get('title') or '').strip()
        if not title:
            return
        db.execute('''INSERT INTO opportunities (title, amount, stage, status, customer_id, lead_id, notes, close_date, created_at, updated_at)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                   (title, to_float(row.get('amount')), row.get('stage') or 'prospecting', row.get('status') or 'open', row.get('customer_id') or None, row.get('lead_id') or None, row.get('notes'), row.get('close_date') or None, now, now))
        count += 1

    if filename.endswith('.csv'):
        data = file.read().decode('utf-8')
        reader = csv.DictReader(StringIO(data))
        for row in reader:
            insert_from(row)
    elif filename.endswith(('.xlsx', '.xls')):
        if not openpyxl:
            flash('openpyxl not installed on server')
            return redirect(url_for('opportunities_list'))
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            insert_from({headers[i]: row[i] for i in range(len(headers))})
    else:
        flash('Unsupported file format. Use CSV or XLSX.')
        return redirect(url_for('opportunities_list'))

    db.commit()
    flash(f'Imported {count} opportunities')
    return redirect(url_for('opportunities_list'))

@app.route('/opportunities/new', methods=['GET', 'POST'])
@login_required
def opportunities_new():
    db = get_db()
    customers = db.execute('SELECT id, name FROM customers ORDER BY name').fetchall()
    leads = db.execute('SELECT id, name FROM leads ORDER BY name').fetchall()
    if request.method == 'POST':
        data = {
            'title': request.form.get('title', '').strip(),
            'amount': request.form.get('amount') or None,
            'stage': request.form.get('stage', 'prospecting'),
            'status': request.form.get('status', 'open'),
            'customer_id': request.form.get('customer_id') or None,
            'lead_id': request.form.get('lead_id') or None,
            'notes': request.form.get('notes'),
            'close_date': request.form.get('close_date') or None
        }
        if not data['title']:
            flash('Title is required')
            return redirect(url_for('opportunities_new'))
        db.execute(
            '''INSERT INTO opportunities (title, amount, stage, status, customer_id, lead_id, notes, close_date, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (data['title'], data['amount'], data['stage'], data['status'], data['customer_id'], data['lead_id'], data['notes'], data['close_date'], datetime.utcnow().isoformat(), datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Opportunity created')
        return redirect(url_for('opportunities_list'))
    return render_template('opportunities/form.html', opportunity=None, customers=customers, leads=leads)

@app.route('/opportunities/<int:opp_id>/edit', methods=['GET', 'POST'])
@login_required
def opportunities_edit(opp_id):
    db = get_db()
    opportunity = db.execute('SELECT * FROM opportunities WHERE id = ?', (opp_id,)).fetchone()
    customers = db.execute('SELECT id, name FROM customers ORDER BY name').fetchall()
    leads = db.execute('SELECT id, name FROM leads ORDER BY name').fetchall()
    if not opportunity:
        flash('Opportunity not found')
        return redirect(url_for('opportunities_list'))
    if request.method == 'POST':
        data = {
            'title': request.form.get('title', '').strip(),
            'amount': request.form.get('amount') or None,
            'stage': request.form.get('stage', opportunity['stage']),
            'status': request.form.get('status', opportunity['status']),
            'customer_id': request.form.get('customer_id') or None,
            'lead_id': request.form.get('lead_id') or None,
            'notes': request.form.get('notes'),
            'close_date': request.form.get('close_date') or None
        }
        if not data['title']:
            flash('Title is required')
            return redirect(url_for('opportunities_edit', opp_id=opp_id))
        db.execute(
            '''UPDATE opportunities SET title=?, amount=?, stage=?, status=?, customer_id=?, lead_id=?, notes=?, close_date=?, updated_at=? WHERE id=?''',
            (data['title'], data['amount'], data['stage'], data['status'], data['customer_id'], data['lead_id'], data['notes'], data['close_date'], datetime.utcnow().isoformat(), opp_id)
        )
        db.commit()
        flash('Opportunity updated')
        return redirect(url_for('opportunities_list'))
    return render_template('opportunities/form.html', opportunity=opportunity, customers=customers, leads=leads)

@app.route('/opportunities/<int:opp_id>/delete', methods=['POST'])
@login_required
def opportunities_delete(opp_id):
    db = get_db()
    db.execute('DELETE FROM opportunities WHERE id = ?', (opp_id,))
    db.commit()
    flash('Opportunity deleted')
    return redirect(url_for('opportunities_list'))

# ----------------------
# API - Customers
# ----------------------

@app.route('/api/customers', methods=['GET', 'POST'])
@login_required
def api_customers():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute('SELECT * FROM customers ORDER BY updated_at DESC').fetchall()
        return jsonify([dict(r) for r in rows])
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name is required'}), 400
    values = (
        name,
        data.get('email'), data.get('mobile'), data.get('phone'), data.get('address'),
        data.get('city'), data.get('pincode'), data.get('state'), data.get('country'),
        data.get('notes'), data.get('status', 'active'), datetime.utcnow().isoformat(), datetime.utcnow().isoformat()
    )
    cur = db.execute('''INSERT INTO customers (name, email, mobile, phone, address, city, pincode, state, country, notes, status, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', values)
    db.commit()
    new_id = cur.lastrowid
    row = db.execute('SELECT * FROM customers WHERE id=?', (new_id,)).fetchone()
    return jsonify(dict(row)), 201

@app.route('/api/customers/<int:customer_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_customer_detail(customer_id):
    db = get_db()
    row = db.execute('SELECT * FROM customers WHERE id=?', (customer_id,)).fetchone()
    if not row:
        return jsonify({'error': 'not found'}), 404
    if request.method == 'GET':
        return jsonify(dict(row))
    if request.method == 'DELETE':
        db.execute('DELETE FROM customers WHERE id=?', (customer_id,))
        db.commit()
        return jsonify({'status': 'deleted'})
    # PUT
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or row['name']).strip()
    values = (
        name,
        data.get('email', row['email']), data.get('mobile', row['mobile']), data.get('phone', row['phone']),
        data.get('address', row['address']), data.get('city', row['city']), data.get('pincode', row['pincode']),
        data.get('state', row['state']), data.get('country', row['country']), data.get('notes', row['notes']),
        data.get('status', row['status']), datetime.utcnow().isoformat(), customer_id
    )
    db.execute('''UPDATE customers SET name=?, email=?, mobile=?, phone=?, address=?, city=?, pincode=?, state=?, country=?, notes=?, status=?, updated_at=? WHERE id=?''', values)
    db.commit()
    row = db.execute('SELECT * FROM customers WHERE id=?', (customer_id,)).fetchone()
    return jsonify(dict(row))

# ----------------------
# API - Leads
# ----------------------

@app.route('/api/leads', methods=['GET', 'POST'])
@login_required
def api_leads():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute('SELECT * FROM leads ORDER BY updated_at DESC').fetchall()
        return jsonify([dict(r) for r in rows])
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name is required'}), 400
    cur = db.execute('''INSERT INTO leads (name, email, mobile, phone, source, status, notes, customer_id, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (name, data.get('email'), data.get('mobile'), data.get('phone'), data.get('source'), data.get('status', 'new'), data.get('notes'), data.get('customer_id'), datetime.utcnow().isoformat(), datetime.utcnow().isoformat()))
    db.commit()
    row = db.execute('SELECT * FROM leads WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201

@app.route('/api/leads/<int:lead_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_lead_detail(lead_id):
    db = get_db()
    row = db.execute('SELECT * FROM leads WHERE id=?', (lead_id,)).fetchone()
    if not row:
        return jsonify({'error': 'not found'}), 404
    if request.method == 'GET':
        return jsonify(dict(row))
    if request.method == 'DELETE':
        db.execute('DELETE FROM leads WHERE id=?', (lead_id,))
        db.commit()
        return jsonify({'status': 'deleted'})
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or row['name']).strip()
    db.execute('''UPDATE leads SET name=?, email=?, mobile=?, phone=?, source=?, status=?, notes=?, customer_id=?, updated_at=? WHERE id=?''',
               (name, data.get('email', row['email']), data.get('mobile', row['mobile']), data.get('phone', row['phone']), data.get('source', row['source']), data.get('status', row['status']), data.get('notes', row['notes']), data.get('customer_id', row['customer_id']), datetime.utcnow().isoformat(), lead_id))
    db.commit()
    row = db.execute('SELECT * FROM leads WHERE id=?', (lead_id,)).fetchone()
    return jsonify(dict(row))

# ----------------------
# API - Opportunities
# ----------------------

@app.route('/api/opportunities', methods=['GET', 'POST'])
@login_required
def api_opportunities():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute('SELECT * FROM opportunities ORDER BY updated_at DESC').fetchall()
        return jsonify([dict(r) for r in rows])
    data = request.get_json(silent=True) or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({'error': 'title is required'}), 400
    cur = db.execute('''INSERT INTO opportunities (title, amount, stage, status, customer_id, lead_id, notes, close_date, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (title, data.get('amount'), data.get('stage', 'prospecting'), data.get('status', 'open'), data.get('customer_id'), data.get('lead_id'), data.get('notes'), data.get('close_date'), datetime.utcnow().isoformat(), datetime.utcnow().isoformat()))
    db.commit()
    row = db.execute('SELECT * FROM opportunities WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201

@app.route('/api/opportunities/<int:opp_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_opportunity_detail(opp_id):
    db = get_db()
    row = db.execute('SELECT * FROM opportunities WHERE id=?', (opp_id,)).fetchone()
    if not row:
        return jsonify({'error': 'not found'}), 404
    if request.method == 'GET':
        return jsonify(dict(row))
    if request.method == 'DELETE':
        db.execute('DELETE FROM opportunities WHERE id=?', (opp_id,))
        db.commit()
        return jsonify({'status': 'deleted'})
    data = request.get_json(silent=True) or {}
    title = (data.get('title') or row['title']).strip()
    db.execute('''UPDATE opportunities SET title=?, amount=?, stage=?, status=?, customer_id=?, lead_id=?, notes=?, close_date=?, updated_at=? WHERE id=?''',
               (title, data.get('amount', row['amount']), data.get('stage', row['stage']), data.get('status', row['status']), data.get('customer_id', row['customer_id']), data.get('lead_id', row['lead_id']), data.get('notes', row['notes']), data.get('close_date', row['close_date']), datetime.utcnow().isoformat(), opp_id))
    db.commit()
    row = db.execute('SELECT * FROM opportunities WHERE id=?', (opp_id,)).fetchone()
    return jsonify(dict(row))

# ----------------------
# App startup
# ----------------------

with app.app_context():
    init_db()

if __name__ == '__main__':
    # In dev mode, enable debug server
    app.run(debug=True, port=5020)